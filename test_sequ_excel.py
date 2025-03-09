import openpyxl
import pandas as pd


class test_sequence:
    # Load the Excel file
    def __init__(self,file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(self.file_path)

        # test sequence
        self.ws = self.wb["BLP_Hotfire"]  #Get the active sheet
        # Extract and index each cell with row and column numbers
        self.sequence_data = []
        self.abort_data    = []

        self.test_num = 0
        self.duration = 0
        self.s_duration = 0
        self.time      = [] # prepend the rest of the dat
        self.functions = []
        self.actions   = []
        self.Limit     = []
        self.cond      = []
        self.unit      = []
        self.oob       = []

    def parse_test(self):
        index     = 0
        #test_num
        #self.test_num = self.ws.cell(row = 1, column = 2).value
        #print("test_num " + str(test_num))
        self.duration = self.ws.cell(row = 2, column = 2).value
        #print("duration " + str(self.duration))
        self.s_duration = self.ws.cell(row = 3, column = 2).value
        #print("s_duration " + str(self.s_duration))

        #add test duration details to index
        self.sequence_data.append([self.duration])
        self.sequence_data.append([self.s_duration])

        #print(self.sequence_data)
        row_data = []
        for i in range(1,27):
            row_data = []
            for j in range(1,8):
                cell_value = self.ws.cell(row= (i+5), column=j).value  # Get cell value
                row_data.append(cell_value)
            self.sequence_data.append(row_data)
            #print(indexed_data[index])
            index += 1
        #print(self.sequence_data)
        return self.sequence_data

    def parse_abort_limit(self):
        index = 0
        row_data = []
        for i in range(3):
            row_data = []
            for j in range(1,8):
                cell_value = self.ws.cell(row= (i+29), column=j).value  # Get cell value
                row_data.append(cell_value)
            self.abort_data.append(row_data)
            #print(indexed_data[index])
            index += 1
        #print(self.abort_data)
        return self.abort_data



t = test_sequence("Hotfire_Basic_Sample.xlsx")
print(t.parse_test())
t.parse_abort_limit()
      
